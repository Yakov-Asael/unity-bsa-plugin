"""
Compare two quarters of the Salesforce "IPE RawData" export and produce a
highlighted "To Review" workbook.

Usage:
    python compare_quarters.py <prev_quarter_file> <curr_quarter_file> <output_xlsx> \
        [--prev-sheet "RawData"] [--curr-sheet "RawData"] [--key-col Id] [--diff-start-col J]

Input files can be .xlsx or .csv. For .xlsx, the named sheet is read.

Logic:
  - Match rows between quarters by the key column (default: Id).
  - Diff columns are everything from --diff-start-col (default "J") to the last
    column present in the header row. Columns before that (identity fields) are
    never diffed.
  - A row is "changed" if ANY diff column differs between quarters for the same key.
  - A row is "new" if its key exists in curr but not prev.
  - A key is "removed" if it exists in prev but not curr (can't appear as a row in
    the output since it's not in curr — reported separately in the JSON summary).
  - Output workbook has one sheet ("To Review") with curr quarter's data, same
    column order as input, changed rows filled yellow, new rows filled green.
  - Prints a JSON summary to stdout: changed users (with which columns flipped and
    old/new values), new users, removed user ids.
"""
import argparse
import json
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")


def read_rows(path, sheet_name):
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [r for r in reader]
    else:
        wb = load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet '{sheet_name}' not found in {path}. Sheets: {wb.sheetnames}")
        ws = wb[sheet_name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
    header = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    return header, data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("prev_file")
    ap.add_argument("curr_file")
    ap.add_argument("output_xlsx")
    ap.add_argument("--prev-sheet", default="RawData")
    ap.add_argument("--curr-sheet", default="RawData")
    ap.add_argument("--key-col", default="Id")
    ap.add_argument("--diff-start-col", default="J")
    args = ap.parse_args()

    prev_header, prev_rows = read_rows(args.prev_file, args.prev_sheet)
    curr_header, curr_rows = read_rows(args.curr_file, args.curr_sheet)

    if prev_header != curr_header:
        print(json.dumps({
            "error": "header_mismatch",
            "prev_header": prev_header,
            "curr_header": curr_header,
        }))
        sys.exit(1)

    header = curr_header
    key_idx = header.index(args.key_col)
    diff_start_idx = column_index_from_string(args.diff_start_col) - 1
    diff_cols = list(range(diff_start_idx, len(header)))

    prev_by_key = {r[key_idx]: r for r in prev_rows if r[key_idx] is not None}
    curr_by_key = {r[key_idx]: r for r in curr_rows if r[key_idx] is not None}

    changed = {}
    new_keys = []
    for key, row in curr_by_key.items():
        if key not in prev_by_key:
            new_keys.append(key)
            continue
        prev_row = prev_by_key[key]
        diffs = []
        for ci in diff_cols:
            pv = prev_row[ci] if ci < len(prev_row) else None
            cv = row[ci] if ci < len(row) else None
            if str(pv) != str(cv):
                diffs.append({"column": header[ci], "old": pv, "new": cv})
        if diffs:
            changed[key] = diffs

    removed_keys = [k for k in prev_by_key if k not in curr_by_key]

    wb = Workbook()
    ws = wb.active
    ws.title = "To Review"
    ws.append(header)
    name_idx = header.index("Name") if "Name" in header else None
    for row in curr_rows:
        ws.append(row)
        r = ws.max_row
        key = row[key_idx]
        if key in changed:
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).fill = YELLOW
        elif key in new_keys:
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).fill = GREEN
    wb.save(args.output_xlsx)

    def label(key, rows_by_key):
        row = rows_by_key.get(key)
        if row and name_idx is not None and name_idx < len(row):
            return row[name_idx]
        return None

    summary = {
        "changed_users": [
            {"id": k, "name": label(k, curr_by_key), "diffs": v}
            for k, v in changed.items()
        ],
        "new_users": [
            {"id": k, "name": label(k, curr_by_key)} for k in new_keys
        ],
        "removed_users": [
            {"id": k, "name": label(k, prev_by_key)} for k in removed_keys
        ],
    }
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
