"""
Flag genuine permission exceptions in a quarter's PermissionSetAssignment
RawData, using Test 2's Definition tab. Handles the X<ProfileId> profile-
mirror resolution and named-individual approved-list entries. See
references/permset-schema.md and references/permset-definition-rules.md.

Usage:
    python permset_exceptions.py <data_file> <definition_source_file> \
        <profile_lookup_file> <output_xlsx> \
        [--data-sheet "Raw Data"] [--definition-sheet "Definition"] \
        [--profile-sheet "RawData"]

<profile_lookup_file> is Test 1's RawData workbook (or any file with a
Profile.Id / Profile.Name column pair) — used to resolve X<ProfileId> rows
back to a real profile name.
"""
import argparse
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

RED = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
ORANGE = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

SSO_COLUMN = "PermissionSet.PermissionsIsSsoEnabled"
X_PROFILE_RE = re.compile(r"^X([0-9A-Za-z]{15,18})$")
X_ID_SHAPED_RE = re.compile(r"^X([0-9A-Za-z]{15,18})")  # not anchored at the end — catches suffixed/mangled variants

NAMED_INDIVIDUAL_HINTS = ("only visable", "only visible", "- approved", " and ")

# Ids the user has manually confirmed the identity of, where automated
# resolution (matching against Test 1's Profile.Id list) isn't possible.
# Confirmed by the user: X00ex00000018ozT_128_09_43_34_1 is a real, distinct
# "System Administrator Custom" permission set — NOT a plain profile-mirror
# of the System Administrator profile. Per the user's instruction, this
# identity is approved ONLY on the specific columns where the Definition tab
# names it explicitly ("System Administrator Custom", exact) — it must NOT
# get swept in wherever plain "System Administrator" is approved via the
# normal loose/substring matching, so it's listed here for strict-only
# matching rather than being aliased to "System Administrator" outright.
KNOWN_ID_ALIASES = {
    "00ex00000018ozT": "System Administrator Custom",
}
STRICT_IDENTITIES = {"system administrator custom"}


def build_profile_lookup(path, sheet):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    id_idx = header.index("Profile.Id")
    name_idx = header.index("Profile.Name")
    lookup = {}
    for r in rows[1:]:
        if r[id_idx] is not None:
            lookup[r[id_idx]] = r[name_idx]
    return lookup


def parse_definition(ws):
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for i, r in enumerate(rows):
        if r and r[0] == "Permission name":
            start = i + 1
            break
    if start is None:
        raise SystemExit("Could not find 'Permission name' header in Definition tab")

    rules = {}
    for r in rows[start:]:
        if not r or r[0] is None:
            continue
        names = [n.strip() for n in str(r[0]).split("\n") if n.strip()]
        approved_raw = r[2] if len(r) > 2 else None
        tokens = [t.strip() for t in str(approved_raw).split("\n") if t.strip()] if approved_raw else []
        for n in names:
            rules[n] = tokens
            # The Definition tab sometimes copy-pastes a "Profile." prefix
            # instead of "PermissionSet." (seen on the SSO row) — register
            # under both so a lookup by the actual data column name succeeds
            # regardless of which prefix the Definition tab used.
            if n.startswith("Profile."):
                rules[n.replace("Profile.", "PermissionSet.", 1)] = tokens
            elif n.startswith("PermissionSet."):
                rules[n.replace("PermissionSet.", "Profile.", 1)] = tokens
    return rules


def is_named_individual_list(tokens):
    joined = " ".join(tokens).lower()
    return any(hint in joined for hint in NAMED_INDIVIDUAL_HINTS)


def _normalize(s):
    s = s.lower().strip()
    s = re.sub(r"^permission set( called)?[\s\-:]*", "", s)
    s = s.replace("_", " ")
    s = re.sub(r"\(.*?\)", "", s)  # drop parenthetical annotations
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip("s")  # crude pluralization handling (Users -> User)
    return s


def identity_is_approved(identity, tokens):
    if any(t.lower() == "all" for t in tokens):
        return True
    norm_identity = _normalize(identity) if identity else ""
    strict = norm_identity in STRICT_IDENTITIES
    for t in tokens:
        if t.lower() == "n/a":
            continue
        norm_t = _normalize(t)
        if identity == t or norm_identity == norm_t:
            return True
        if strict:
            continue  # exact-normalized match only; no loose/prefix fallback
        # profile-name-with-parenthetical-suffix match (exact prefix, unnormalized)
        if t.startswith(identity) and (len(t) == len(identity) or t[len(identity)] in " ("):
            return True
        # loose containment either direction, on normalized forms
        if norm_identity and (norm_identity in norm_t or norm_t in norm_identity):
            return True
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("data_file")
    ap.add_argument("definition_source_file")
    ap.add_argument("profile_lookup_file")
    ap.add_argument("output_xlsx")
    ap.add_argument("--data-sheet", default="Raw Data")
    ap.add_argument("--definition-sheet", default="Definition")
    ap.add_argument("--profile-sheet", default="RawData")
    args = ap.parse_args()

    data_wb = load_workbook(args.data_file, data_only=True)
    data_ws = data_wb[args.data_sheet]
    rows = list(data_ws.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]

    def_wb = load_workbook(args.definition_source_file, data_only=True)
    def_ws = def_wb[args.definition_sheet]
    rules = parse_definition(def_ws)

    profile_lookup = build_profile_lookup(args.profile_lookup_file, args.profile_sheet)

    uname_idx = header.index("Assignee.Username")
    pset_idx = header.index("PermissionSet.Name")
    perm_start = header.index(SSO_COLUMN)
    perm_cols = header[perm_start:]

    exceptions = []
    manual_review = []
    definition_updates = {}  # column -> set of note strings

    for r in data_rows:
        pset_name = r[pset_idx]
        clean_m = X_PROFILE_RE.match(pset_name) if pset_name else None
        shaped_m = X_ID_SHAPED_RE.match(pset_name) if pset_name else None
        resolved_profile = profile_lookup.get(shaped_m.group(1)) if shaped_m else None
        known_alias = KNOWN_ID_ALIASES.get(shaped_m.group(1)) if shaped_m else None
        if resolved_profile:
            resolved_identity = resolved_profile
            source_note = f"profile-mirror of {resolved_identity}"
            unresolved = False
        elif known_alias:
            resolved_identity = known_alias
            source_note = f"user-confirmed identity: {known_alias}"
            unresolved = False
        else:
            resolved_identity = pset_name
            unresolved = bool(shaped_m)  # looks ID-shaped (clean or suffixed) but didn't resolve
            source_note = ("unresolved profile-mirror-shaped Id (no matching Profile.Id found)"
                            if unresolved else "explicit permission set")

        for i, col in enumerate(perm_cols):
            ci = perm_start + i
            val = r[ci] if ci < len(r) else None
            tokens = rules.get(col, [])

            is_true = val is True or str(val).lower() == "true"
            is_sso_bad = col == SSO_COLUMN and (val is False or str(val).lower() == "false")
            triggers = is_true if col != SSO_COLUMN else is_sso_bad

            if not triggers:
                continue

            if is_named_individual_list(tokens):
                manual_review.append({
                    "Assignee.Username": r[uname_idx],
                    "PermissionSet.Name": pset_name,
                    "Resolved": resolved_identity,
                    "Permission": col,
                    "Value": val,
                    "Reason": "approved list names specific individuals — verify identity manually",
                })
                continue

            if unresolved:
                manual_review.append({
                    "Assignee.Username": r[uname_idx],
                    "PermissionSet.Name": pset_name,
                    "Resolved": resolved_identity,
                    "Permission": col,
                    "Value": val,
                    "Reason": "PermissionSet.Name looks like a profile-mirror Id but doesn't match any known Profile.Id — identity unresolved, don't auto-flag as a confirmed exception",
                })
                continue

            approved = identity_is_approved(resolved_identity, tokens)
            if not approved:
                exceptions.append({
                    "Assignee.Username": r[uname_idx],
                    "PermissionSet.Name": pset_name,
                    "Resolved": resolved_identity,
                    "Source": source_note,
                    "Permission": col,
                    "Value": val,
                })
                note = (f"flagged — {r[uname_idx]} ({resolved_identity}) has this permission "
                        f"TRUE via {pset_name}, not on approved list. Needs review.")
                definition_updates.setdefault(col, set()).add(note)

    wb = Workbook()
    ws = wb.active
    ws.title = "Exceptions"
    ws.append(["Assignee.Username", "PermissionSet.Name", "Resolved", "Source", "Permission", "Value"])
    for e in exceptions:
        ws.append([e["Assignee.Username"], e["PermissionSet.Name"], e["Resolved"], e["Source"], e["Permission"], e["Value"]])
        for c in range(1, 7):
            ws.cell(row=ws.max_row, column=c).fill = RED

    ws2 = wb.create_sheet("Manual Review Needed")
    ws2.append(["Assignee.Username", "PermissionSet.Name", "Resolved", "Permission", "Value", "Reason"])
    for e in manual_review:
        ws2.append([e["Assignee.Username"], e["PermissionSet.Name"], e["Resolved"], e["Permission"], e["Value"], e["Reason"]])
        for c in range(1, 7):
            ws2.cell(row=ws2.max_row, column=c).fill = ORANGE

    def_copy = wb.create_sheet("Definition")
    for row in def_ws.iter_rows(values_only=True):
        def_copy.append(row)

    wb.save(args.output_xlsx)

    print(json.dumps({
        "exception_count": len(exceptions),
        "manual_review_count": len(manual_review),
        "exceptions": exceptions,
        "manual_review": manual_review,
        "definition_e_column_updates": {k: sorted(v) for k, v in definition_updates.items()},
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
