"""
Flag genuine permission exceptions in a quarter's RawData, using the
Definition tab's per-permission approved-profile lists to exclude expected
combinations. See references/definition-rules.md for the parsing/logic spec.

Usage:
    python flag_exceptions.py <data_file> <definition_source_file> <output_xlsx> \
        [--data-sheet "Sheet1"] [--definition-sheet "Definition"] [--key-col Id]

<data_file> is the quarter's raw data (any sheet name, via --data-sheet).
<definition_source_file> is a workbook containing the `Definition` tab
(usually the master tracking workbook, since bare per-quarter exports don't
carry it) — can be the same file as <data_file> if it happens to have one.
"""
import argparse
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

RED = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

SSO_COLUMN = "Profile.PermissionsIsSsoEnabled"
BLANKET_APPROVED_PROFILE = "System Administrator"


def parse_definition(ws):
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for i, r in enumerate(rows):
        if r and r[0] == "Permission name":
            start = i + 1
            break
    if start is None:
        raise SystemExit("Could not find 'Permission name' header in Definition tab")

    rules = {}  # column_name -> set of approved profiles, or "ALL"
    for r in rows[start:]:
        if not r or r[0] is None:
            continue
        names = [n.strip() for n in str(r[0]).split("\n") if n.strip()]
        approved_raw = r[2] if len(r) > 2 else None
        if approved_raw is None:
            approved = set()
        else:
            tokens = [t.strip() for t in str(approved_raw).split("\n") if t.strip()]
            if any(t.lower() == "all" for t in tokens):
                approved = "ALL"
            else:
                approved = set(t for t in tokens if t.lower() != "n/a")
        for n in names:
            rules[n] = approved
    return rules


def profile_is_approved(profile, approved):
    if approved == "ALL":
        return True
    if profile in approved:
        return True
    # Definition tab sometimes appends a parenthetical description to a
    # profile name (e.g. "Sales Insights Integration User (For Sales
    # integration)") that doesn't match Salesforce's actual Profile.Name
    # field exactly. Treat it as approved if the approved-list entry starts
    # with the actual profile name followed by a space/paren.
    for a in approved:
        if a.startswith(profile) and (len(a) == len(profile) or a[len(profile)] in " ("):
            return True
    return False


def is_exception(column, value, profile, rules):
    approved = rules.get(column, set())
    if profile == BLANKET_APPROVED_PROFILE:
        return False
    if column == SSO_COLUMN:
        # default expectation TRUE; exception if FALSE and not approved
        return (value is False or str(value).lower() == "false") and not profile_is_approved(profile, approved)
    # default expectation FALSE; exception if TRUE and not approved
    return (value is True or str(value).lower() == "true") and not profile_is_approved(profile, approved)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("data_file")
    ap.add_argument("definition_source_file")
    ap.add_argument("output_xlsx")
    ap.add_argument("--data-sheet", default="Sheet1")
    ap.add_argument("--definition-sheet", default="Definition")
    ap.add_argument("--key-col", default="Id")
    args = ap.parse_args()

    data_wb = load_workbook(args.data_file, data_only=True)
    data_ws = data_wb[args.data_sheet]
    rows = list(data_ws.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = [r for r in rows[1:] if r[header.index(args.key_col)] is not None]

    def_wb = load_workbook(args.definition_source_file, data_only=True)
    def_ws = def_wb[args.definition_sheet]
    rules = parse_definition(def_ws)

    id_idx = header.index("Id")
    name_idx = header.index("Name")
    profile_idx = header.index("Profile.Name")

    perm_start = header.index("Profile.PermissionsIsSsoEnabled")
    perm_cols = header[perm_start:]

    exceptions = []
    for r in data_rows:
        profile = r[profile_idx]
        for i, col in enumerate(perm_cols):
            ci = perm_start + i
            val = r[ci] if ci < len(r) else None
            if is_exception(col, val, profile, rules):
                exceptions.append({
                    "Id": r[id_idx],
                    "Name": r[name_idx],
                    "Profile.Name": profile,
                    "Permission": col,
                    "Value": val,
                })

    wb = Workbook()
    ws = wb.active
    ws.title = "Exceptions"
    ws.append(["Id", "Name", "Profile.Name", "Permission", "Value"])
    for e in exceptions:
        ws.append([e["Id"], e["Name"], e["Profile.Name"], e["Permission"], e["Value"]])
        for c in range(1, 6):
            ws.cell(row=ws.max_row, column=c).fill = RED

    # Copy the Definition tab in verbatim so the workbook is self-contained
    def_copy = wb.create_sheet("Definition")
    for row in def_ws.iter_rows(values_only=True):
        def_copy.append(row)

    wb.save(args.output_xlsx)
    print(json.dumps({"exception_count": len(exceptions), "exceptions": exceptions}, indent=2, default=str))


if __name__ == "__main__":
    main()
