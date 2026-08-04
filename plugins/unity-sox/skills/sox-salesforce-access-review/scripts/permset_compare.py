"""
Compare two quarters of PermissionSetAssignment RawData using the composite
key (Assignee.Username, PermissionSet.Name), since there's no single unique
record Id in this export. See references/permset-schema.md.

Usage:
    python permset_compare.py <prev_file> <curr_file> <output_json> \
        [--prev-sheet "Raw Data"] [--curr-sheet "Raw Data"]

Prints a JSON summary: new assignment rows, removed assignment rows, and
rows where a permission value changed for the same (Assignee, PermissionSet)
pair. Does not produce a highlighted workbook itself — pair with
permset_exceptions.py for the actual red-cell/exception output, since that's
what matters most for this test.
"""
import argparse
import json
from openpyxl import load_workbook


def read_rows(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {path}. Sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    header = [str(h) if h is not None else "" for h in rows[0]]
    data = [r for r in rows[1:] if any(v is not None for v in r)]
    return header, data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("prev_file")
    ap.add_argument("curr_file")
    ap.add_argument("output_json")
    ap.add_argument("--prev-sheet", default="Raw Data")
    ap.add_argument("--curr-sheet", default="Raw Data")
    args = ap.parse_args()

    prev_header, prev_rows = read_rows(args.prev_file, args.prev_sheet)
    curr_header, curr_rows = read_rows(args.curr_file, args.curr_sheet)

    if prev_header != curr_header:
        print(json.dumps({"error": "header_mismatch", "prev_header": prev_header, "curr_header": curr_header}))
        raise SystemExit(1)

    header = curr_header
    uname_idx = header.index("Assignee.Username")
    pset_idx = header.index("PermissionSet.Name")
    perm_start = header.index("PermissionSet.PermissionsIsSsoEnabled")

    def key(r):
        return (r[uname_idx], r[pset_idx])

    prev_by_key = {key(r): r for r in prev_rows if r[uname_idx] is not None}
    curr_by_key = {key(r): r for r in curr_rows if r[uname_idx] is not None}

    new_keys = [k for k in curr_by_key if k not in prev_by_key]
    removed_keys = [k for k in prev_by_key if k not in curr_by_key]
    changed = {}
    for k, row in curr_by_key.items():
        if k not in prev_by_key:
            continue
        prev_row = prev_by_key[k]
        diffs = []
        for ci in range(perm_start, len(header)):
            pv = prev_row[ci] if ci < len(prev_row) else None
            cv = row[ci] if ci < len(row) else None
            if str(pv) != str(cv):
                diffs.append({"column": header[ci], "old": pv, "new": cv})
        if diffs:
            changed[f"{k[0]} | {k[1]}"] = diffs

    summary = {
        "new_assignments": [{"username": k[0], "permission_set": k[1]} for k in new_keys],
        "removed_assignments": [{"username": k[0], "permission_set": k[1]} for k in removed_keys],
        "changed_assignments": changed,
    }
    with open(args.output_json, "w") as f:
        json.dump(summary, f, indent=2, default=str)
    print(json.dumps({
        "new_count": len(new_keys),
        "removed_count": len(removed_keys),
        "changed_count": len(changed),
    }, indent=2))


if __name__ == "__main__":
    main()
