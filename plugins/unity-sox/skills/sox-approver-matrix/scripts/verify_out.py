#!/usr/bin/env python
"""Independent check of the finished workbooks, after build.py has written them."""
import glob, os, sys, warnings, zipfile
warnings.filterwarnings("ignore")
import openpyxl
from common import OUT, objects

EXPECT_TABS = ["Results", "IPE", "Modify All Permissions", "Custom Settings", "Approval"]

ok = True
found = 0
for f in sorted(glob.glob(os.path.join(OUT, "*.xlsx"))):
    if os.path.basename(f).startswith("~$"):     # Excel lock file
        continue
    found += 1
    wb = openpyxl.load_workbook(f)
    media = [n for n in zipfile.ZipFile(f).namelist() if n.startswith("xl/media/")]
    res = wb["Results"] if "Results" in wb.sheetnames else None
    ipe = wb["IPE"] if "IPE" in wb.sheetnames else None
    appr = wb["Approval"] if "Approval" in wb.sheetnames else None

    verdict = []
    if ipe:
        verdict = [ipe.cell(r, 1).value for r in range(1, ipe.max_row + 1)
                   if isinstance(ipe.cell(r, 1).value, str) and "MATCH" in ipe.cell(r, 1).value]
    approvers = [appr.cell(r, 1).value for r in range(2, appr.max_row + 1)] if appr else []

    tabs_ok = wb.sheetnames == EXPECT_TABS
    print(os.path.basename(f))
    print(f"   tabs      : {wb.sheetnames}  {'OK' if tabs_ok else '<-- MISMATCH'}")
    print(f"   Results   : {res.max_row - 1 if res else '?'} data rows, "
          f"{res.max_column if res else '?'} cols")
    print(f"   images    : {len(media)} embedded")
    print(f"   IPE check : {verdict}")
    print(f"   approvers : {approvers}")

    if not tabs_ok or not verdict or any("MISMATCH" in v for v in verdict):
        ok = False
    if len(media) < 5:
        print("   <-- too few screenshots embedded")
        ok = False
    if not approvers:
        print("   <-- no approvers listed")
        ok = False

if found != len(objects()):
    print(f"\nExpected {len(objects())} workbooks, found {found}")
    ok = False

print("\nALL CHECKS PASSED" if ok else "\nCHECKS FAILED")
sys.exit(0 if ok else 1)
