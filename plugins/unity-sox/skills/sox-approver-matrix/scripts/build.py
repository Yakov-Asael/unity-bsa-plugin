#!/usr/bin/env python
"""
Builds the quarterly SOX approver-matrix workbooks.

Input : data/<key>.json (from convert.py), shots/*.png, timings.json
Output: ../output/<Object> Approver Matrix - <Month> <Year>.xlsx
"""
import datetime, json, os, sys
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from common import HERE, SHOTS, DATA, OUT, objects, permission_sets, load_json

os.makedirs(OUT, exist_ok=True)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True)
LBL_FONT = Font(bold=True, size=12)
MONO = Font(name="Menlo", size=10)
IMG_W = 1180          # embedded width in px; keeps the clock legible, files sane


def put_image(ws, path, anchor):
    """Embed a screenshot at `anchor`, scaled to IMG_W, and return rows consumed."""
    if not os.path.exists(path):
        ws[anchor] = f"[MISSING SCREENSHOT: {os.path.basename(path)}]"
        ws[anchor].font = Font(color="C00000", bold=True)
        return 2
    with PILImage.open(path) as im:
        w, h = im.size
    img = XLImage(path)
    img.width, img.height = IMG_W, int(h * IMG_W / w)
    ws.add_image(img, anchor)
    return int(img.height / 19) + 3          # ~19px per default row


def sheet_results(wb, obj):
    ws = wb.create_sheet("Results")
    cols, rows = obj["columns"], obj["rows"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    for i, col in enumerate(cols, 1):
        width = max([len(str(col))] + [len(str(r[i - 1] or "")) for r in rows]) if rows else len(col)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 55)
    ws.freeze_panes = "A2"
    return len(rows)


def sheet_ipe(wb, key, total_records, row_count, timing):
    ws = wb.create_sheet("IPE")
    ws.column_dimensions["A"].width = 120
    r = 1
    ws.cell(r, 1, "Report run — Total Records").font = LBL_FONT
    r += 2
    r += put_image(ws, os.path.join(SHOTS, f"{key}_ipe_run.png"), f"A{r}")
    ws.cell(r, 1, "Report filters").font = LBL_FONT
    r += 2
    r += put_image(ws, os.path.join(SHOTS, f"{key}_ipe_filters.png"), f"A{r}")

    ws.cell(r, 1, "Completeness & accuracy check").font = LBL_FONT
    r += 1
    if timing:
        fmt = "%d/%m/%Y %H:%M:%S"
        ws.cell(r, 1, "Screenshot taken:              "
                + datetime.datetime.fromtimestamp(timing["shot"]).strftime(fmt)); r += 1
        ws.cell(r, 1, "Report exported to 'Results':  "
                + datetime.datetime.fromtimestamp(timing["csv"]).strftime(fmt)); r += 1
        ws.cell(r, 1, f"Gap:                           {timing['csv'] - timing['shot']} seconds"); r += 1
    ws.cell(r, 1, f"Total Records shown in report: {total_records}"); r += 1
    ws.cell(r, 1, f"Rows in 'Results' tab:         {row_count}"); r += 1
    ok = (total_records == row_count)
    c = ws.cell(r, 1, "MATCH" if ok else "MISMATCH — investigate before submitting")
    c.font = Font(bold=True, color="006100" if ok else "C00000")
    return ok


def sheet_modify_all(wb, key, obj):
    ws = wb.create_sheet("Modify All Permissions")
    ws.column_dimensions["A"].width = 120
    r = 1
    ws.cell(r, 1, f"Profiles / Permission Sets with Modify All on {obj['sobject']}").font = LBL_FONT
    r += 2
    for line in obj["objperm_query"].split("\n"):
        ws.cell(r, 1, line).font = MONO
        r += 1
    r += 1
    r += put_image(ws, os.path.join(SHOTS, f"{key}_modifyall.png"), f"A{r}")

    for i, h in enumerate(["Parent.Name", "Parent.Profile.Name", "SobjectType",
                           "ModifyAllRecords"], 1):
        cell = ws.cell(r, i, h); cell.fill, cell.font = HDR_FILL, HDR_FONT
    r += 1
    for rec in obj["objperm_rows"]:
        for i, v in enumerate(rec, 1):
            ws.cell(r, i, v)
        r += 1
    r += 2

    for ps in obj["permission_sets"]:
        ws.cell(r, 1, f"Permission Set: {ps['name']}").font = LBL_FONT
        r += 1
        for line in ps["query"].split("\n"):
            ws.cell(r, 1, line).font = MONO
            r += 1
        r += 1
        r += put_image(ws, os.path.join(SHOTS, f"{key}_ps_{ps['name']}.png"), f"A{r}")
        for i, h in enumerate(["Assignee.Name", "Assignee.Profile.Name",
                               "Assignee.FN_Role__c"], 1):
            cell = ws.cell(r, i, h); cell.fill, cell.font = HDR_FILL, HDR_FONT
        r += 1
        for rec in ps["assignments"]:
            for i, v in enumerate(rec, 1):
                ws.cell(r, i, v)
            r += 1
        ws.cell(r, 1, f"Total assigned: {len(ps['assignments'])}").font = Font(bold=True)
        r += 3
    for col, w in zip("BCD", (28, 26, 22)):
        ws.column_dimensions[col].width = w


def sheet_custom_settings(wb):
    ws = wb.create_sheet("Custom Settings")
    ws.column_dimensions["A"].width = 120
    r = 1
    ws.cell(r, 1, "Custom Setting definition (incl. Last Modified Date)").font = LBL_FONT
    r += 2
    r += put_image(ws, os.path.join(SHOTS, "cs_definition.png"), f"A{r}")
    ws.cell(r, 1, "Custom Setting rows — all records").font = LBL_FONT
    r += 2
    put_image(ws, os.path.join(SHOTS, "cs_rows.png"), f"A{r}")


def sheet_approval(wb, approvers):
    ws = wb.create_sheet("Approval")
    hdr = ["Approver Name", "Approval Date", "Approved? (Y/N)", "Comment"]
    ws.append(hdr)
    for i in range(1, len(hdr) + 1):
        cell = ws.cell(1, i); cell.fill, cell.font = HDR_FILL, HDR_FONT
    for name in approvers:
        ws.append([name, "", "", ""])
    for col, w in zip("ABCD", (28, 16, 18, 60)):
        ws.column_dimensions[col].width = w


def main():
    now = datetime.datetime.now()
    stamp = f"{now.strftime('%B')} {now.year}"
    try:
        timings = load_json(os.path.join(HERE, "timings.json"))
    except FileNotFoundError:
        timings = {}

    summary = []
    for o in objects():
        key = o["key"]
        obj = load_json(os.path.join(DATA, f"{key}.json"))
        wb = Workbook(); wb.remove(wb.active)
        n = sheet_results(wb, obj)
        ok = sheet_ipe(wb, key, obj["total_records"], n, timings.get(key))
        sheet_modify_all(wb, key, obj)
        sheet_custom_settings(wb)
        sheet_approval(wb, obj.get("approvers", []))
        name = f"{obj['display']} Approver Matrix - {stamp}.xlsx"
        wb.save(os.path.join(OUT, name))
        summary.append((name, n, obj["total_records"], ok))

    print(f"{'file':52s} {'rows':>6s} {'report':>7s}  check")
    for name, n, tr, ok in summary:
        print(f"{name:52s} {n:6d} {tr:7d}  {'OK' if ok else 'MISMATCH'}")
    if not all(s[3] for s in summary):
        sys.exit(1)


if __name__ == "__main__":
    main()
