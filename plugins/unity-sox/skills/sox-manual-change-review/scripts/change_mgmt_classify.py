"""
Classify a quarter's Setup Audit Trail export for the Change Management SOX
review: exclude listed users, filter to the target quarter by month, look up
Type from Section, and flag Relevant/Manual Review using the whitelist
approach (see references/change-mgmt-relevance-rules.md — do NOT use a
Type-based or generic-keyword rule, both tested badly against real data).

Usage:
    python change_mgmt_classify.py <raw_data_file> <exclude_users_file> \
        <section_type_file> <whitelist_file> <output_xlsx> \
        --quarter Q2 [--data-sheet "Sheet1"] [--exclude-sheet "Users to Exclude"] \
        [--section-type-sheet "Section -> Type"]

<raw_data_file>, <exclude_users_file>, <section_type_file> can all be the
same workbook (the master tracking sheet) or separate files — pass sheet
names accordingly. <whitelist_file> is the JSON file in references/.
"""
import argparse
import json
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
ORANGE = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

QUARTER_MONTHS = {
    "Q1": {1, 2, 3},
    "Q2": {4, 5, 6},
    "Q3": {7, 8, 9},
    "Q4": {10, 11, 12},
}

COSMETIC_PATTERNS = ["page layout", "picklist", "changed label of custom field", "custom formula field"]


def extract_name(action):
    action = action or ""
    m = re.search(r'"([^"]+)"', action)
    if m:
        return m.group(1).strip()
    m = re.search(r"Approval Process:\s*(.+)$", action)
    if m:
        return m.group(1).strip()
    return None


def is_cosmetic(action):
    a = (action or "").lower()
    return any(p in a for p in COSMETIC_PATTERNS)


def parse_date_string(s):
    """Parse Salesforce's raw audit-trail date format: 'DD/MM/YYYY HH:MM:SS TZ'
    (e.g. '02/07/2026 14:57:30 IDT'). Returns a datetime or None."""
    if not s:
        return None
    s = str(s).strip()
    parts = s.rsplit(" ", 1)
    candidate = parts[0] if len(parts) == 2 and parts[1].isalpha() else s
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(candidate, fmt)
        except ValueError:
            continue
    return None


def read_rows(path, sheet_name):
    """Load rows from either a raw CSV export or an xlsx workbook tab."""
    if path.lower().endswith(".csv"):
        import csv
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = [tuple(r) for r in csv.reader(f)]
        except UnicodeDecodeError:
            with open(path, newline="", encoding="latin-1") as f:
                rows = [tuple(r) for r in csv.reader(f)]
        return rows
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    return list(ws.iter_rows(values_only=True))


def get_month(row, month_idx, date_idx):
    if month_idx is not None and row[month_idx]:
        try:
            return int(row[month_idx])
        except (TypeError, ValueError):
            pass
    if date_idx is not None and row[date_idx]:
        d = row[date_idx]
        if isinstance(d, datetime):
            return d.month
        parsed = parse_date_string(d)
        if parsed:
            return parsed.month
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("raw_data_file")
    ap.add_argument("exclude_users_file")
    ap.add_argument("section_type_file")
    ap.add_argument("whitelist_file")
    ap.add_argument("output_xlsx")
    ap.add_argument("--quarter", required=True, choices=["Q1", "Q2", "Q3", "Q4"])
    ap.add_argument("--data-sheet", default="Raw Data")
    ap.add_argument("--exclude-sheet", default="Users to Exclude")
    ap.add_argument("--section-type-sheet", default="Section -> Type")
    args = ap.parse_args()

    target_months = QUARTER_MONTHS[args.quarter]

    excl_wb = load_workbook(args.exclude_users_file, data_only=True, read_only=True)
    excl_ws = excl_wb[args.exclude_sheet]
    exclude_users = set()
    for r in excl_ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            exclude_users.add(str(r[0]).strip())

    st_wb = load_workbook(args.section_type_file, data_only=True, read_only=True)
    st_ws = st_wb[args.section_type_sheet]
    section_to_type = {}
    for r in st_ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            section_to_type[str(r[0]).strip()] = r[1].strip() if r[1] else None

    whitelist = set(json.load(open(args.whitelist_file))["whitelist"])

    data_ws_rows = read_rows(args.raw_data_file, args.data_sheet)
    rows = data_ws_rows
    header = list(rows[0])

    date_idx = header.index("Date") if "Date" in header else None
    month_idx = header.index("Month") if "Month" in header else None
    user_idx = header.index("User")
    action_idx = header.index("Action")
    section_idx = header.index("Section")

    results = []
    for r in rows[1:]:
        if not any(v is not None for v in r):
            continue
        if len(r) <= max(user_idx, action_idx, section_idx, date_idx or 0, month_idx or 0):
            continue  # short/ragged row, skip rather than crash
        month = get_month(r, month_idx, date_idx)
        if month not in target_months:
            continue
        user = r[user_idx]
        if user in exclude_users:
            continue
        section = r[section_idx]
        action = r[action_idx]
        type_ = section_to_type.get(section, "UNKNOWN SECTION — not in Section -> Type tab")

        if is_cosmetic(action):
            relevant = "Manual Review"
        else:
            name = extract_name(action)
            relevant = "Confident Relevant" if (name and name in whitelist) else "Manual Review"

        results.append({
            "row": list(r),
            "type": type_,
            "relevant": relevant,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = f"{args.quarter} Data Filtered"
    out_header = list(header) + ["Type", "Relevant"]
    for j, val in enumerate(out_header, start=1):
        ws.cell(row=1, column=j, value=val)
    for i, res in enumerate(results, start=2):
        fill = GREEN if res["relevant"] == "Confident Relevant" else ORANGE
        out_row = res["row"] + [res["type"], res["relevant"]]
        for j, val in enumerate(out_row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = fill

    review_ws = wb.create_sheet(f"{args.quarter} Changes to Review")
    for j, val in enumerate(out_header, start=1):
        review_ws.cell(row=1, column=j, value=val)
    ri = 2
    for res in results:
        if res["relevant"] == "Confident Relevant":
            out_row = res["row"] + [res["type"], res["relevant"]]
            for j, val in enumerate(out_row, start=1):
                review_ws.cell(row=ri, column=j, value=val)
            ri += 1

    wb.save(args.output_xlsx)

    counts = {"Confident Relevant": 0, "Manual Review": 0}
    unknown_sections = set()
    for res in results:
        counts[res["relevant"]] += 1
        if res["type"] and res["type"].startswith("UNKNOWN"):
            unknown_sections.add(res["row"][section_idx])

    print(json.dumps({
        "quarter": args.quarter,
        "total_rows_in_scope": len(results),
        "counts": counts,
        "unknown_sections_needing_manual_type_lookup": sorted(s for s in unknown_sections if s is not None),
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
